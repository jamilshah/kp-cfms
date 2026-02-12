"""
-------------------------------------------------------------------------
System: KP-CFMS (Computerized Financial Management System)
Client: Local Government Department, Khyber Pakhtunkhwa
Team Lead: Jamil Shah
Developers: Ali Asghar, Akhtar Munir and Zarif Khan
Description: Views for the Revenue module including Payer, Demand,
             and Collection management with workflow support.
-------------------------------------------------------------------------
"""
from typing import Any, Dict
import json
import csv
from decimal import Decimal
from django.views.generic import (
    ListView, CreateView, DetailView, UpdateView, TemplateView, View
)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy, reverse
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Sum, Count, Q, Max
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_protect
from django.utils.translation import gettext_lazy as _
from django.utils import timezone

from apps.users.permissions import MakerRequiredMixin, CashierRequiredMixin, CheckerRequiredMixin
from apps.revenue.models import (
    Payer, RevenueDemand, RevenueCollection, 
    DemandStatus, CollectionStatus
)
from apps.revenue.forms import (
    PayerForm, DemandForm, CollectionForm,
    DemandPostForm, CollectionPostForm, DemandCancelForm
)
from apps.revenue.reports import RevenueReports
from apps.budgeting.models import FiscalYear, Department
from apps.finance.models import BudgetHead, AccountType, FunctionCode
from apps.core.exceptions import WorkflowTransitionException


class RevenueDashboardView(LoginRequiredMixin, TemplateView):
    """
    Dashboard view for the revenue module.
    
    Shows summary statistics for demands and collections.
    """
    template_name = 'revenue/dashboard.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = getattr(user, 'organization', None)
        
        # Get current operating fiscal year (based on today's date)
        active_fy = FiscalYear.get_current_operating_year() if org else None
        
        if org and active_fy:
            demands_qs = RevenueDemand.objects.filter(
                organization=org,
                fiscal_year=active_fy
            )
            
            collections_qs = RevenueCollection.objects.filter(
                organization=org,
                demand__fiscal_year=active_fy
            )
            
            # Demand statistics
            total_demands = demands_qs.exclude(status=DemandStatus.CANCELLED).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            
            total_collections = collections_qs.filter(
                status=CollectionStatus.POSTED
            ).aggregate(
                total=Sum('amount_received')
            )['total'] or Decimal('0.00')
            
            context.update({
                'active_fiscal_year': active_fy,
                'total_demands': total_demands,
                'total_collections': total_collections,
                'outstanding_balance': total_demands - total_collections,
                'demand_count': demands_qs.exclude(status=DemandStatus.CANCELLED).count(),
                'draft_demands': demands_qs.filter(status=DemandStatus.DRAFT).count(),
                'posted_demands': demands_qs.filter(status=DemandStatus.POSTED).count(),
                'partial_demands': demands_qs.filter(status=DemandStatus.PARTIAL).count(),
                'paid_demands': demands_qs.filter(status=DemandStatus.PAID).count(),
                'collection_count': collections_qs.filter(status=CollectionStatus.POSTED).count(),
                'pending_collections': collections_qs.filter(status=CollectionStatus.DRAFT),
                'recent_demands': demands_qs.order_by('-created_at')[:5],
                'recent_collections': collections_qs.filter(
                    status=CollectionStatus.POSTED
                ).order_by('-receipt_date')[:5],
            })
        
        return context


# ============================================================================
# Payer Views
# ============================================================================

class PayerListView(LoginRequiredMixin, ListView):
    """List view for payers."""
    
    model = Payer
    template_name = 'revenue/payer_list.html'
    context_object_name = 'payers'
    paginate_by = 20
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        if org:
            queryset = Payer.objects.filter(organization=org).order_by('name')
            
            # Search by name or CNIC
            search = self.request.GET.get('search', '')
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) | Q(cnic_ntn__icontains=search)
                )
            
            return queryset
        return Payer.objects.none()
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class PayerCreateView(LoginRequiredMixin, MakerRequiredMixin, CreateView):
    """Create view for payers.
    
    Requires Dealing Assistant (Maker) role.
    """
    
    model = Payer
    form_class = PayerForm
    template_name = 'revenue/payer_form.html'
    success_url = reverse_lazy('revenue:payer_list')
    
    
    def get_template_names(self):
        if self.request.headers.get('HX-Request'):
            return ['revenue/payer_form_partial.html']
        return [self.template_name]

    def form_valid(self, form):
        user = self.request.user
        org = getattr(user, 'organization', None)
        form.instance.organization = org
        form.instance.created_by = user
        
        if self.request.headers.get('HX-Request'):
            self.object = form.save()
            response = HttpResponse(status=204)
            trigger_data = {
                'payerCreated': {
                    'id': self.object.id,
                    'name': self.object.name
                }
            }
            response['HX-Trigger'] = json.dumps(trigger_data)
            return response
            
        messages.success(self.request, _('Payer created successfully.'))
        return super().form_valid(form)


class PayerUpdateView(LoginRequiredMixin, UpdateView):
    """Update view for payers."""
    
    model = Payer
    form_class = PayerForm
    template_name = 'revenue/payer_form.html'
    success_url = reverse_lazy('revenue:payer_list')
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        if org:
            return Payer.objects.filter(organization=org)
        return Payer.objects.none()
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Payer updated successfully.'))
        return super().form_valid(form)


class PayerDetailView(LoginRequiredMixin, DetailView):
    """Detail view for payers with demand history."""
    
    model = Payer
    template_name = 'revenue/payer_detail.html'
    context_object_name = 'payer'
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        if org:
            return Payer.objects.filter(organization=org)
        return Payer.objects.none()
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['demands'] = self.object.demands.order_by('-issue_date')[:10]
        context['total_demands'] = self.object.get_total_demands()
        context['total_collected'] = self.object.get_total_collections()
        context['outstanding'] = self.object.get_outstanding_balance()
        return context


# ============================================================================
# Demand Views
# ============================================================================

class DemandListView(LoginRequiredMixin, ListView):
    """List view for revenue demands."""
    
    model = RevenueDemand
    template_name = 'revenue/demand_list.html'
    context_object_name = 'demands'
    paginate_by = 20
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        
        if not org:
            return RevenueDemand.objects.none()
        
        queryset = RevenueDemand.objects.filter(
            organization=org
        ).select_related(
            'payer',
            'budget_head',
            'budget_head__nam_head',
            'budget_head__function',
            'fiscal_year',
            'posted_by',
            'cancelled_by',
            'accrual_voucher'
        ).prefetch_related(
            'collections'
        ).order_by('-issue_date', '-created_at')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status in dict(DemandStatus.choices):
            queryset = queryset.filter(status=status)
        
        # Filter by fiscal year
        fy_id = self.request.GET.get('fiscal_year')
        if fy_id:
            queryset = queryset.filter(fiscal_year_id=fy_id)
        
        # Search by challan no or payer name
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(challan_no__icontains=search) | Q(payer__name__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = getattr(user, 'organization', None)
        
        context['status_choices'] = DemandStatus.choices
        context['current_status'] = self.request.GET.get('status', '')
        context['search_query'] = self.request.GET.get('search', '')
        
        if org:
            context['fiscal_years'] = FiscalYear.objects.all().order_by('-start_date')
        
        return context


class DemandCreateView(LoginRequiredMixin, MakerRequiredMixin, CreateView):
    """Create view for revenue demands.
    
    Requires Dealing Assistant (Maker) role.
    """
    
    model = RevenueDemand
    form_class = DemandForm
    template_name = 'revenue/demand_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        """Check for current operating fiscal year before allowing access."""
        user = request.user
        org = getattr(user, 'organization', None)
        
        if org:
            active_fy = FiscalYear.get_current_operating_year()
            if not active_fy:
                messages.error(
                    request, 
                    _('No active fiscal year found for the current period. '
                      'Please contact your administrator to set up a fiscal year.')
                )
                return redirect('revenue:dashboard')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        user = self.request.user
        kwargs['organization'] = getattr(user, 'organization', None)
        return kwargs
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = getattr(user, 'organization', None)
        if org:
            context['current_fiscal_year'] = FiscalYear.get_current_operating_year()
        return context
    
    def form_valid(self, form):
        user = self.request.user
        org = getattr(user, 'organization', None)
        
        form.instance.organization = org
        form.instance.created_by = user
        
        # Set fiscal year from form's cached value
        if form._current_fiscal_year:
            form.instance.fiscal_year = form._current_fiscal_year
        
        # Generate challan number with transaction lock to prevent race conditions
        with transaction.atomic():
            fy = form.instance.fiscal_year
            
            # Get the max challan number for this fiscal year with lock
            last_demand = RevenueDemand.objects.select_for_update().filter(
                organization=org,
                fiscal_year=fy
            ).aggregate(Max('challan_no'))
            
            # Extract counter from last challan number
            if last_demand['challan_no__max']:
                # Parse CH-2025-26-0001 → get 0001 → increment
                parts = last_demand['challan_no__max'].split('-')
                try:
                    counter = int(parts[-1]) + 1
                except (ValueError, IndexError):
                    counter = 1
            else:
                counter = 1
            
            form.instance.challan_no = f"CH-{fy.year_name}-{counter:04d}"
            
            messages.success(self.request, _('Demand created successfully.'))
            return super().form_valid(form)
    
    def get_success_url(self) -> str:
        return reverse('revenue:demand_detail', kwargs={'pk': self.object.pk})


class DemandDetailView(LoginRequiredMixin, DetailView):
    """Detail view for demands with collection history."""
    
    model = RevenueDemand
    template_name = 'revenue/demand_detail.html'
    context_object_name = 'demand'
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        if org:
            return RevenueDemand.objects.filter(
                organization=org
            ).select_related(
                'payer', 'budget_head', 'fiscal_year', 
                'accrual_voucher', 'posted_by'
            )
        return RevenueDemand.objects.none()
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['collections'] = self.object.collections.order_by('-receipt_date')
        context['total_collected'] = self.object.get_total_collected()
        context['outstanding'] = self.object.get_outstanding_balance()
        context['can_post'] = self.object.status == DemandStatus.DRAFT
        context['can_collect'] = self.object.status in [
            DemandStatus.POSTED, DemandStatus.PARTIAL
        ]
        context['can_cancel'] = self.object.status in [
            DemandStatus.DRAFT, DemandStatus.POSTED
        ] and not self.object.collections.filter(
            status=CollectionStatus.POSTED
        ).exists()
        # Can waive penalty if: penalty enabled, not yet waived, and demand is overdue
        context['can_waive_penalty'] = (
            self.object.apply_penalty and 
            not self.object.penalty_waived and
            self.object.status in [DemandStatus.POSTED, DemandStatus.PARTIAL] and
            self.object.get_days_overdue() > 0
        )
        return context


class DemandPostView(LoginRequiredMixin, CheckerRequiredMixin, View):
    """View for posting a demand to GL.
    
    Requires Finance Officer/Checker role to post demands.
    """
    
    def post(self, request, pk):
        user = request.user
        org = getattr(user, 'organization', None)
        
        demand = get_object_or_404(
            RevenueDemand.objects.select_for_update().filter(organization=org),
            pk=pk
        )
        
        # Idempotency check
        if demand.status != DemandStatus.DRAFT:
            messages.warning(request, _('Demand has already been posted.'))
            return redirect('revenue:demand_detail', pk=pk)
        
        try:
            demand.post(user)
            messages.success(
                request,
                _(f'Demand {demand.challan_no} posted successfully. '
                  f'Voucher {demand.accrual_voucher.voucher_no} created.')
            )
        except WorkflowTransitionException as e:
            messages.error(request, f'Workflow Error: {str(e)}')
        except ValidationError as e:
            error_msg = ', '.join([f"{k}: {v}" for k, v in e.message_dict.items()])
            messages.error(request, f'Validation Error: {error_msg}')
        except Exception as e:
            messages.error(request, _('Error posting demand: ') + str(e))
        
        return redirect('revenue:demand_detail', pk=pk)


class DemandCancelView(LoginRequiredMixin, MakerRequiredMixin, View):
    """View for cancelling a demand.
    
    Requires Maker role or higher to cancel demands.
    """
    
    def post(self, request, pk):
        user = request.user
        org = getattr(user, 'organization', None)
        
        demand = get_object_or_404(
            RevenueDemand.objects.select_for_update().filter(organization=org),
            pk=pk
        )
        
        reason = request.POST.get('reason', '')
        if not reason:
            messages.error(request, _('Cancellation reason is required.'))
            return redirect('revenue:demand_detail', pk=pk)
        
        try:
            demand.cancel(user, reason)
            messages.success(request, _('Demand cancelled successfully.'))
        except WorkflowTransitionException as e:
            messages.error(request, f'Workflow Error: {str(e)}')
        except ValidationError as e:
            error_msg = ', '.join([f"{k}: {v}" for k, v in e.message_dict.items()])
            messages.error(request, f'Validation Error: {error_msg}')
        except Exception as e:
            messages.error(request, _('Error cancelling demand: ') + str(e))
        
        return redirect('revenue:demand_detail', pk=pk)


class DemandWaivePenaltyView(LoginRequiredMixin, CheckerRequiredMixin, View):
    """View for waiving penalty on a demand.
    
    Requires Checker/Finance Officer role to waive penalties.
    """
    
    def post(self, request, pk):
        user = request.user
        org = getattr(user, 'organization', None)
        
        demand = get_object_or_404(
            RevenueDemand.objects.select_for_update().filter(organization=org),
            pk=pk
        )
        
        # Get waiver reason
        reason = request.POST.get('reason', '').strip()
        if not reason:
            messages.error(request, _('Waiver reason is required.'))
            return redirect('revenue:demand_detail', pk=pk)
        
        try:
            demand.waive_penalty(user, reason)
            messages.success(
                request, 
                _(f'Penalty for demand {demand.challan_no} has been waived.')
            )
        except ValidationError as e:
            if hasattr(e, 'message'):
                messages.error(request, e.message)
            else:
                messages.error(request, str(e))
        except Exception as e:
            messages.error(request, _('Error waiving penalty: ') + str(e))
        
        return redirect('revenue:demand_detail', pk=pk)


# ============================================================================
# Collection Views
# ============================================================================

class CollectionListView(LoginRequiredMixin, ListView):
    """List view for revenue collections."""
    
    model = RevenueCollection
    template_name = 'revenue/collection_list.html'
    context_object_name = 'collections'
    paginate_by = 20
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        
        if not org:
            return RevenueCollection.objects.none()
        
        queryset = RevenueCollection.objects.filter(
            organization=org
        ).select_related(
            'demand',
            'demand__payer',
            'demand__budget_head',
            'demand__budget_head__nam_head',
            'demand__fiscal_year',
            'bank_account',
            'bank_account__gl_code',
            'posted_by',
            'receipt_voucher'
        ).order_by('-receipt_date', '-created_at')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status in dict(CollectionStatus.choices):
            queryset = queryset.filter(status=status)
        
        # Search by receipt no or payer name
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(receipt_no__icontains=search) | 
                Q(demand__payer__name__icontains=search) |
                Q(demand__challan_no__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['status_choices'] = CollectionStatus.choices
        context['current_status'] = self.request.GET.get('status', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class CollectionCreateView(LoginRequiredMixin, CashierRequiredMixin, CreateView):
    """Create view for revenue collections.
    
    Requires Cashier role to record cash/bank receipts.
    """
    
    model = RevenueCollection
    form_class = CollectionForm
    template_name = 'revenue/collection_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        """Check for current operating fiscal year."""
        user = request.user
        org = getattr(user, 'organization', None)
        
        if org:
            active_fy = FiscalYear.get_current_operating_year()
            if not active_fy:
                messages.error(
                    request, 
                    _('No active fiscal year found for the current period.')
                )
                return redirect('revenue:dashboard')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self) -> Dict[str, Any]:
        kwargs = super().get_form_kwargs()
        user = self.request.user
        kwargs['organization'] = getattr(user, 'organization', None)
        
        # Pre-select demand if provided in URL
        demand_id = self.request.GET.get('demand')
        if demand_id:
            try:
                demand = RevenueDemand.objects.get(pk=demand_id)
                kwargs['demand'] = demand
            except RevenueDemand.DoesNotExist:
                pass
        
        return kwargs
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        
        # If demand is pre-selected, add to context
        demand_id = self.request.GET.get('demand')
        if demand_id:
            try:
                demand = RevenueDemand.objects.get(pk=demand_id)
                context['selected_demand'] = demand
                context['outstanding_balance'] = demand.get_outstanding_balance()
            except RevenueDemand.DoesNotExist:
                pass
        
        return context
    
    def form_valid(self, form):
        user = self.request.user
        org = getattr(user, 'organization', None)
        
        form.instance.organization = org
        form.instance.created_by = user
        
        # Generate receipt number with transaction lock to prevent race conditions
        with transaction.atomic():
            fy = form.cleaned_data['demand'].fiscal_year
            
            # Get the max receipt number for this fiscal year with lock
            last_receipt = RevenueCollection.objects.select_for_update().filter(
                organization=org,
                demand__fiscal_year=fy
            ).aggregate(Max('receipt_no'))
            
            # Extract counter from last receipt number
            if last_receipt['receipt_no__max']:
                # Parse REC-2025-26-0001 → get 0001 → increment
                parts = last_receipt['receipt_no__max'].split('-')
                try:
                    counter = int(parts[-1]) + 1
                except (ValueError, IndexError):
                    counter = 1
            else:
                counter = 1
            
            form.instance.receipt_no = f"REC-{fy.year_name}-{counter:04d}"
            
            messages.success(self.request, _('Collection recorded successfully.'))
            return super().form_valid(form)
    
    def get_success_url(self) -> str:
        return reverse('revenue:collection_detail', kwargs={'pk': self.object.pk})


class CollectionDetailView(LoginRequiredMixin, DetailView):
    """Detail view for collections."""
    
    model = RevenueCollection
    template_name = 'revenue/collection_detail.html'
    context_object_name = 'collection'
    
    def get_queryset(self):
        user = self.request.user
        org = getattr(user, 'organization', None)
        if org:
            return RevenueCollection.objects.filter(
                organization=org
            ).select_related(
                'demand', 'demand__payer', 'bank_account',
                'receipt_voucher', 'posted_by'
            )
        return RevenueCollection.objects.none()
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['can_post'] = self.object.status == CollectionStatus.DRAFT
        context['can_cancel'] = self.object.status in [
            CollectionStatus.DRAFT, CollectionStatus.POSTED
        ]
        return context


class CollectionPostView(LoginRequiredMixin, CheckerRequiredMixin, View):
    """View for posting a collection to GL.
    
    Requires Finance Officer/Checker role to post collections.
    """
    
    def post(self, request, pk):
        user = request.user
        org = getattr(user, 'organization', None)
        
        collection = get_object_or_404(
            RevenueCollection.objects.select_for_update().filter(organization=org),
            pk=pk
        )
        
        # Idempotency check
        if collection.status != CollectionStatus.DRAFT:
            messages.warning(request, _('Collection has already been posted.'))
            return redirect('revenue:collection_detail', pk=pk)
        
        try:
            collection.post(user)
            messages.success(
                request,
                _(f'Collection {collection.receipt_no} posted successfully. '
                  f'Voucher {collection.receipt_voucher.voucher_no} created.')
            )
        except WorkflowTransitionException as e:
            messages.error(request, f'Workflow Error: {str(e)}')
        except ValidationError as e:
            error_msg = ', '.join([f"{k}: {v}" for k, v in e.message_dict.items()])
            messages.error(request, f'Validation Error: {error_msg}')
        except Exception as e:
            messages.error(request, _('Error posting collection: ') + str(e))
        
        return redirect('revenue:collection_detail', pk=pk)


class CollectionCancelView(LoginRequiredMixin, CheckerRequiredMixin, View):
    """View for cancelling a collection.
    
    Requires Finance Officer/Checker role to cancel collections.
    """
    
    def post(self, request, pk):
        user = request.user
        org = getattr(user, 'organization', None)
        
        collection = get_object_or_404(
            RevenueCollection.objects.select_for_update().filter(organization=org),
            pk=pk
        )
        
        try:
            collection.cancel(user)
            messages.success(request, _('Collection cancelled successfully.'))
        except WorkflowTransitionException as e:
            messages.error(request, f'Workflow Error: {str(e)}')
        except ValidationError as e:
            error_msg = ', '.join([f"{k}: {v}" for k, v in e.message_dict.items()])
            messages.error(request, f'Validation Error: {error_msg}')
        except Exception as e:
            messages.error(request, _('Error cancelling collection: ') + str(e))
        
        return redirect('revenue:collection_detail', pk=pk)


# ============================================================================
# API Views
# ============================================================================

class DemandOutstandingAPIView(LoginRequiredMixin, View):
    """API endpoint to get outstanding balance for a demand."""
    
    def get(self, request, pk):
        user = request.user
        org = getattr(user, 'organization', None)
        
        try:
            demand = RevenueDemand.objects.get(pk=pk, organization=org)
            return JsonResponse({
                'demand_amount': str(demand.amount),
                'total_collected': str(demand.get_total_collected()),
                'outstanding': str(demand.get_outstanding_balance()),
                'payer_name': demand.payer.name,
                'challan_no': demand.challan_no,
            })
        except RevenueDemand.DoesNotExist:
            return JsonResponse({'error': 'Demand not found'}, status=404)


@csrf_protect
def load_functions(request):
    """
    AJAX view to load functions filtered by department.
    Returns HTML with function options including default and usage notes.
    """
    department_id = request.GET.get('department')
    functions = FunctionCode.objects.none()
    default_id = None
    
    if department_id:
        try:
            # Explicit integer validation to prevent SQL injection
            dept_id = int(department_id)
            dept = Department.objects.get(id=dept_id)
            functions = dept.related_functions.all().order_by('code')
            default_id = dept.default_function_id if dept.default_function else None
        except (ValueError, TypeError, Department.DoesNotExist):
            pass  # Return empty queryset
            
    # Reuse expenditure partial as it is generic and compatible
    return render(request, 'expenditure/partials/function_options.html', {
        'functions': functions,
        'default_id': default_id
    })


@csrf_protect
def load_revenue_heads(request):
    """
    AJAX view to load revenue budget heads filtered by department and/or function.
    
    NEW BEHAVIOR (after CoA restructuring):
    - Filters by department FK directly on BudgetHead
    - When both department AND function provided: returns precise filtered set
    """
    department_id = request.GET.get('department')
    function_id = request.GET.get('function')
    
    # Base queryset
    budget_heads = BudgetHead.objects.filter(
        nam_head__account_type=AccountType.REVENUE,
        posting_allowed=True,
        is_active=True
    ).select_related('nam_head', 'function', 'department').order_by('nam_head__name', 'nam_head__code')
    
    # PRIMARY FILTER: Use both department + function for most precise results
    if department_id and function_id:
        try:
            dept_id = int(department_id)
            func_id = int(function_id)
            budget_heads = budget_heads.filter(
                department_id=dept_id,
                function_id=func_id
            )
        except (ValueError, TypeError):
            pass
    # Filter by function only
    elif function_id:
        try:
            func_id = int(function_id)
            budget_heads = budget_heads.filter(function_id=func_id)
        except (ValueError, TypeError):
            pass
    # Filter by department only
    elif department_id:
        try:
            dept_id = int(department_id)
            budget_heads = budget_heads.filter(department_id=dept_id)
        except (ValueError, TypeError, Department.DoesNotExist):
            pass
            
    # Reuse expenditure partial as it is generic (uses 'budget_heads' context)
    return render(request, 'expenditure/partials/budget_head_options.html', {'budget_heads': budget_heads})


# =============================================================================
# EXPORT VIEWS (CSV/Excel)
# =============================================================================

class DemandExportView(LoginRequiredMixin, View):
    """
    Export demands to CSV or Excel format.
    
    Supports filtering by:
    - Fiscal year
    - Date range
    - Status
    - Payer
    - Budget head
    """
    
    def get(self, request: HttpRequest) -> HttpResponse:
        """Handle export request."""
        org = getattr(request.user, 'organization', None)
        if not org:
            messages.error(request, _('No organization found for user.'))
            return redirect('revenue:demand_list')
        
        export_format = request.GET.get('format', 'csv')
        
        # Build queryset with filters
        demands = RevenueDemand.objects.filter(
            organization=org
        ).select_related(
            'payer',
            'budget_head',
            'budget_head__nam_head',
            'fiscal_year',
            'posted_by'
        ).order_by('-issue_date')
        
        # Apply filters
        fiscal_year_id = request.GET.get('fiscal_year')
        if fiscal_year_id:
            demands = demands.filter(fiscal_year_id=fiscal_year_id)
        
        status = request.GET.get('status')
        if status:
            demands = demands.filter(status=status)
        
        payer_id = request.GET.get('payer')
        if payer_id:
            demands = demands.filter(payer_id=payer_id)
        
        if export_format == 'excel':
            return self._export_excel(demands, org)
        else:
            return self._export_csv(demands, org)
    
    def _export_csv(self, demands, organization) -> HttpResponse:
        """Export demands to CSV format."""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'revenue_demands_{organization.name}_{timestamp}.csv'
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add UTF-8 BOM for Excel compatibility
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Header row
        writer.writerow([
            'Challan No',
            'Fiscal Year',
            'Payer Name',
            'CNIC/NTN',
            'Revenue Head Code',
            'Revenue Head Name',
            'Issue Date',
            'Due Date',
            'Amount (Rs)',
            'Total Collected (Rs)',
            'Outstanding (Rs)',
            'Status',
            'Period Description',
            'Posted Date',
            'Posted By'
        ])
        
        # Data rows
        for demand in demands:
            writer.writerow([
                demand.challan_no,
                demand.fiscal_year.year_name,
                demand.payer.name,
                demand.payer.cnic_ntn or '',
                demand.budget_head.nam_head.code,
                demand.budget_head.nam_head.name,
                demand.issue_date.strftime('%Y-%m-%d'),
                demand.due_date.strftime('%Y-%m-%d'),
                float(demand.amount),
                float(demand.get_total_collected()),
                float(demand.get_outstanding_balance()),
                demand.get_status_display(),
                demand.period_description or '',
                demand.posted_at.strftime('%Y-%m-%d %H:%M') if demand.posted_at else '',
                demand.posted_by.get_full_name() if demand.posted_by else ''
            ])
        
        return response
    
    def _export_excel(self, demands, organization) -> HttpResponse:
        """Export demands to Excel format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messages.error(
                self.request,
                _('Excel export requires openpyxl. Please use CSV format.')
            )
            return redirect('revenue:demand_list')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Revenue Demands"
        
        # Header row with formatting
        headers = [
            'Challan No', 'Fiscal Year', 'Payer Name', 'CNIC/NTN',
            'Revenue Head Code', 'Revenue Head Name', 'Issue Date', 'Due Date',
            'Amount (Rs)', 'Total Collected (Rs)', 'Outstanding (Rs)',
            'Status', 'Period Description', 'Posted Date', 'Posted By'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, demand in enumerate(demands, 2):
            ws.cell(row=row_num, column=1).value = demand.challan_no
            ws.cell(row=row_num, column=2).value = demand.fiscal_year.year_name
            ws.cell(row=row_num, column=3).value = demand.payer.name
            ws.cell(row=row_num, column=4).value = demand.payer.cnic_ntn or ''
            ws.cell(row=row_num, column=5).value = demand.budget_head.nam_head.code
            ws.cell(row=row_num, column=6).value = demand.budget_head.nam_head.name
            ws.cell(row=row_num, column=7).value = demand.issue_date
            ws.cell(row=row_num, column=8).value = demand.due_date
            ws.cell(row=row_num, column=9).value = float(demand.amount)
            ws.cell(row=row_num, column=10).value = float(demand.get_total_collected())
            ws.cell(row=row_num, column=11).value = float(demand.get_outstanding_balance())
            ws.cell(row=row_num, column=12).value = demand.get_status_display()
            ws.cell(row=row_num, column=13).value = demand.period_description or ''
            ws.cell(row=row_num, column=14).value = (
                demand.posted_at.strftime('%Y-%m-%d %H:%M') if demand.posted_at else ''
            )
            ws.cell(row=row_num, column=15).value = (
                demand.posted_by.get_full_name() if demand.posted_by else ''
            )
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Prepare response
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'revenue_demands_{organization.name}_{timestamp}.xlsx'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class CollectionExportView(LoginRequiredMixin, View):
    """
    Export collections to CSV or Excel format.
    
    Supports filtering by:
    - Fiscal year
    - Date range
    - Status
    - Bank account
    """
    
    def get(self, request: HttpRequest) -> HttpResponse:
        """Handle export request."""
        org = getattr(request.user, 'organization', None)
        if not org:
            messages.error(request, _('No organization found for user.'))
            return redirect('revenue:collection_list')
        
        export_format = request.GET.get('format', 'csv')
        
        # Build queryset with filters
        collections = RevenueCollection.objects.filter(
            organization=org
        ).select_related(
            'demand',
            'demand__payer',
            'demand__budget_head',
            'demand__budget_head__nam_head',
            'demand__fiscal_year',
            'bank_account',
            'posted_by'
        ).order_by('-receipt_date')
        
        # Apply filters
        fiscal_year_id = request.GET.get('fiscal_year')
        if fiscal_year_id:
            collections = collections.filter(demand__fiscal_year_id=fiscal_year_id)
        
        status = request.GET.get('status')
        if status:
            collections = collections.filter(status=status)
        
        bank_account_id = request.GET.get('bank_account')
        if bank_account_id:
            collections = collections.filter(bank_account_id=bank_account_id)
        
        if export_format == 'excel':
            return self._export_excel(collections, org)
        else:
            return self._export_csv(collections, org)
    
    def _export_csv(self, collections, organization) -> HttpResponse:
        """Export collections to CSV format."""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'revenue_collections_{organization.name}_{timestamp}.csv'
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add UTF-8 BOM
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Header row
        writer.writerow([
            'Receipt No',
            'Receipt Date',
            'Challan No',
            'Payer Name',
            'Revenue Head',
            'Bank Account',
            'Instrument Type',
            'Instrument No',
            'Amount Received (Rs)',
            'Status',
            'Posted Date',
            'Posted By',
            'Remarks'
        ])
        
        # Data rows
        for collection in collections:
            writer.writerow([
                collection.receipt_no,
                collection.receipt_date.strftime('%Y-%m-%d'),
                collection.demand.challan_no,
                collection.demand.payer.name,
                collection.demand.budget_head.nam_head.name,
                collection.bank_account.title,
                collection.get_instrument_type_display(),
                collection.instrument_no or '',
                float(collection.amount_received),
                collection.get_status_display(),
                collection.posted_at.strftime('%Y-%m-%d %H:%M') if collection.posted_at else '',
                collection.posted_by.get_full_name() if collection.posted_by else '',
                collection.remarks or ''
            ])
        
        return response
    
    def _export_excel(self, collections, organization) -> HttpResponse:
        """Export collections to Excel format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messages.error(
                self.request,
                _('Excel export requires openpyxl. Please use CSV format.')
            )
            return redirect('revenue:collection_list')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Revenue Collections"
        
        # Header row with formatting
        headers = [
            'Receipt No', 'Receipt Date', 'Challan No', 'Payer Name',
            'Revenue Head', 'Bank Account', 'Instrument Type', 'Instrument No',
            'Amount Received (Rs)', 'Status', 'Posted Date', 'Posted By', 'Remarks'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, collection in enumerate(collections, 2):
            ws.cell(row=row_num, column=1).value = collection.receipt_no
            ws.cell(row=row_num, column=2).value = collection.receipt_date
            ws.cell(row=row_num, column=3).value = collection.demand.challan_no
            ws.cell(row=row_num, column=4).value = collection.demand.payer.name
            ws.cell(row=row_num, column=5).value = collection.demand.budget_head.nam_head.name
            ws.cell(row=row_num, column=6).value = collection.bank_account.title
            ws.cell(row=row_num, column=7).value = collection.get_instrument_type_display()
            ws.cell(row=row_num, column=8).value = collection.instrument_no or ''
            ws.cell(row=row_num, column=9).value = float(collection.amount_received)
            ws.cell(row=row_num, column=10).value = collection.get_status_display()
            ws.cell(row=row_num, column=11).value = (
                collection.posted_at.strftime('%Y-%m-%d %H:%M') if collection.posted_at else ''
            )
            ws.cell(row=row_num, column=12).value = (
                collection.posted_by.get_full_name() if collection.posted_by else ''
            )
            ws.cell(row=row_num, column=13).value = collection.remarks or ''
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Prepare response
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'revenue_collections_{organization.name}_{timestamp}.xlsx'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


# =============================================================================
# REPORT VIEWS
# =============================================================================

class OutstandingReceivablesReportView(LoginRequiredMixin, TemplateView):
    """
    Aging analysis report for outstanding receivables.
    
    Shows demands categorized by age: 0-30, 31-60, 61-90, 90+ days
    """
    template_name = 'revenue/reports/outstanding_receivables.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        org = getattr(self.request.user, 'organization', None)
        
        if not org:
            return context
        
        # Get fiscal year from request or use current
        fiscal_year_id = self.request.GET.get('fiscal_year')
        if fiscal_year_id:
            fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        else:
            fiscal_year = FiscalYear.get_current_operating_year()
        
        # Generate aging report
        report_data = RevenueReports.outstanding_receivables_aging(
            organization=org,
            fiscal_year=fiscal_year
        )
        
        context.update({
            'report_data': report_data,
            'fiscal_years': FiscalYear.objects.all().order_by('-start_date'),
            'selected_fiscal_year': fiscal_year
        })
        
        return context


class RevenueByHeadReportView(LoginRequiredMixin, TemplateView):
    """
    Revenue summary by budget head.
    
    Shows demanded, collected, and outstanding amounts for each revenue head.
    """
    template_name = 'revenue/reports/revenue_by_head.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        org = getattr(self.request.user, 'organization', None)
        
        if not org:
            return context
        
        # Get fiscal year from request or use current
        fiscal_year_id = self.request.GET.get('fiscal_year')
        if fiscal_year_id:
            fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        else:
            fiscal_year = FiscalYear.get_current_operating_year()
        
        if not fiscal_year:
            return context
        
        # Generate report
        report_data = RevenueReports.revenue_by_head_summary(
            organization=org,
            fiscal_year=fiscal_year
        )
        
        # Calculate totals
        totals = {
            'total_demanded': sum(item['total_demanded'] for item in report_data),
            'demand_count': sum(item['demand_count'] for item in report_data),
            'total_collected': sum(item['total_collected'] for item in report_data),
            'collection_count': sum(item['collection_count'] for item in report_data),
            'outstanding': sum(item['outstanding'] for item in report_data)
        }
        
        context.update({
            'report_data': report_data,
            'totals': totals,
            'fiscal_years': FiscalYear.objects.all().order_by('-start_date'),
            'selected_fiscal_year': fiscal_year
        })
        
        return context


class PayerWiseSummaryView(LoginRequiredMixin, TemplateView):
    """
    Payer-wise revenue summary.
    
    Shows total demands, collections, and outstanding for each payer.
    """
    template_name = 'revenue/reports/payer_wise_summary.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        org = getattr(self.request.user, 'organization', None)
        
        if not org:
            return context
        
        # Get fiscal year from request
        fiscal_year_id = self.request.GET.get('fiscal_year')
        fiscal_year = None
        if fiscal_year_id:
            fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        
        # Generate report
        report_data = RevenueReports.payer_wise_summary(
            organization=org,
            fiscal_year=fiscal_year
        )
        
        context.update({
            'report_data': report_data,
            'fiscal_years': FiscalYear.objects.all().order_by('-start_date'),
            'selected_fiscal_year': fiscal_year
        })
        
        return context


class CollectionEfficiencyReportView(LoginRequiredMixin, TemplateView):
    """
    Collection efficiency metrics report.
    
    Shows collection rate, average days to collect, on-time vs late payments.
    """
    template_name = 'revenue/reports/collection_efficiency.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        org = getattr(self.request.user, 'organization', None)
        
        if not org:
            return context
        
        # Get fiscal year from request or use current
        fiscal_year_id = self.request.GET.get('fiscal_year')
        if fiscal_year_id:
            fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        else:
            fiscal_year = FiscalYear.get_current_operating_year()
        
        if not fiscal_year:
            return context
        
        # Generate report
        report_data = RevenueReports.collection_efficiency_report(
            organization=org,
            fiscal_year=fiscal_year
        )
        
        context.update({
            'report_data': report_data,
            'fiscal_years': FiscalYear.objects.all().order_by('-start_date'),
            'selected_fiscal_year': fiscal_year
        })
        
        return context


class OverdueDemandsReportView(LoginRequiredMixin, TemplateView):
    """
    Overdue demands report.
    
    Lists all demands past their due date with outstanding balances.
    """
    template_name = 'revenue/reports/overdue_demands.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        org = getattr(self.request.user, 'organization', None)
        
        if not org:
            return context
        
        # Generate report
        report_data = RevenueReports.overdue_demands_report(organization=org)
        
        context.update({
            'report_data': report_data
        })
        
        return context


class DemandVsCollectionComparisonView(LoginRequiredMixin, TemplateView):
    """
    Demand vs Collection comparison report.
    
    Compares demands issued vs collections received by month or budget head.
    """
    template_name = 'revenue/reports/demand_vs_collection.html'
    
    def get_context_data(self, **kwargs) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        org = getattr(self.request.user, 'organization', None)
        
        if not org:
            return context
        
        # Get fiscal year from request or use current
        fiscal_year_id = self.request.GET.get('fiscal_year')
        if fiscal_year_id:
            fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        else:
            fiscal_year = FiscalYear.get_current_operating_year()
        
        if not fiscal_year:
            return context
        
        # Get grouping method
        group_by = self.request.GET.get('group_by', 'month')
        
        # Generate report
        report_data = RevenueReports.demand_vs_collection_comparison(
            organization=org,
            fiscal_year=fiscal_year,
            group_by=group_by
        )
        
        context.update({
            'report_data': report_data,
            'fiscal_years': FiscalYear.objects.all().order_by('-start_date'),
            'selected_fiscal_year': fiscal_year,
            'group_by': group_by
        })
        
        return context
