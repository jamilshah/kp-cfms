# apps/finance/management/commands/import_tma_coa.py

"""
TMA Chart of Accounts Importer

This script imports the official TMA Chart of Accounts from Excel file
with proper hierarchical structure:

Structure:
- Major Head (e.g., A01) - GLOBAL HEAD
  └─ Minor Head (e.g., A011) - GLOBAL HEAD
     └─ Detailed Head (e.g., A01101) - TRANSACTION HEAD
        └─ Sub-Head (e.g., C03880-10) - TRANSACTION HEAD (where applicable)

Key Features:
1. Creates proper 3-level hierarchy (Major > Minor > Detailed)
2. Handles special sub-head sections (e.g., C03880-xx)
3. Sets correct control account flags
4. Maps to account types (ASSET/LIABILITY/REVENUE/EXPENDITURE)
5. Validates data before import
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from apps.finance.models import globalhead, minor_head, detailed_head, account_type
import openpyxl
import logging

logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Import TMA Chart of Accounts from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Path to TMA CoA Excel file'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without saving to database'
        )
        parser.add_argument(
            '--clear-existing',
            action='store_true',
            help='WARNING: Delete all existing Chart of Accounts before import'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        dry_run = options['dry_run']
        clear_existing = options.get('clear_existing', False)

        self.stdout.write(self.style.WARNING('=' * 80))
        self.stdout.write(self.style.WARNING('TMA CHART OF ACCOUNTS IMPORTER'))
        self.stdout.write(self.style.WARNING('=' * 80))
        
        if dry_run:
            self.stdout.write(self.style.NOTICE('\n⚠️  DRY RUN MODE - No changes will be saved\n'))
        
        if clear_existing and not dry_run:
            self.stdout.write(self.style.ERROR(
                '\n⚠️  WARNING: This will DELETE all existing Chart of Accounts!'
            ))
            response = input('Type "DELETE ALL" to confirm: ')
            if response != 'DELETE ALL':
                self.stdout.write('Aborted.')
                return
            
            count = ChartOfAccounts.objects.count()
            ChartOfAccounts.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'✓ Deleted {count} existing accounts\n'))
        
        try:
            # Step 1: Load Excel file
            self.stdout.write('Loading Excel file...')
            accounts_data = self._load_excel(excel_file)
            self.stdout.write(self.style.SUCCESS(f'✓ Loaded {len(accounts_data)} account entries\n'))
            
            # Step 2: Validate data
            self.stdout.write('Validating data...')
            validation_errors = self._validate_data(accounts_data)
            
            if validation_errors:
                self.stdout.write(self.style.ERROR('\n✗ VALIDATION ERRORS FOUND:'))
                for error in validation_errors:
                    self.stdout.write(self.style.ERROR(f'  • {error}'))
                self.stdout.write('\nPlease fix these errors in the Excel file and try again.')
                return
            
            self.stdout.write(self.style.SUCCESS('✓ Data validation passed\n'))
            
            # Step 3: Build hierarchical structure
            self.stdout.write('Building account hierarchy...')
            hierarchy = self._build_hierarchy(accounts_data)
            self._display_hierarchy_summary(hierarchy)
            
            # Step 4: Import to database
            if not dry_run:
                with transaction.atomic():
                    stats = self._import_accounts(hierarchy)
                    self.stdout.write(self.style.SUCCESS('\n✓ Import completed successfully!'))
                    self._display_import_stats(stats)
            else:
                self._preview_import(hierarchy)
            
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'\n✗ File not found: {excel_file}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\n✗ Error: {str(e)}'))
            logger.exception('TMA CoA import failed')
            raise

    def _load_excel(self, filepath):
        """Load account data from Excel file"""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        accounts = []
        
        # Start from row 3 (after headers in row 1-2)
        for row in range(3, ws.max_row + 1):
            major_code = ws.cell(row, 1).value
            major_desc = ws.cell(row, 2).value
            minor_code = ws.cell(row, 3).value
            minor_desc = ws.cell(row, 4).value
            detailed_code = ws.cell(row, 5).value
            detailed_desc = ws.cell(row, 6).value
            
            # Skip completely empty rows
            if not any([major_code, minor_code, detailed_code]):
                continue
            
            accounts.append({
                'row': row,
                'major_code': str(major_code).strip() if major_code else '',
                'major_desc': str(major_desc).strip() if major_desc else '',
                'minor_code': str(minor_code).strip() if minor_code else '',
                'minor_desc': str(minor_desc).strip() if minor_desc else '',
                'detailed_code': str(detailed_code).strip() if detailed_code else '',
                'detailed_desc': str(detailed_desc).strip() if detailed_desc else ''
            })
        
        wb.close()
        return accounts

    def _validate_data(self, accounts_data):
        """Validate loaded data for common issues"""
        errors = []
        
        seen_codes = set()
        
        for acc in accounts_data:
            row = acc['row']
            
            # Check for duplicate codes
            if acc['detailed_code']:
                if acc['detailed_code'] in seen_codes:
                    errors.append(f"Row {row}: Duplicate code '{acc['detailed_code']}'")
                seen_codes.add(acc['detailed_code'])
            
            # Check if detailed code exists without description
            if acc['detailed_code'] and not acc['detailed_desc']:
                errors.append(f"Row {row}: Code '{acc['detailed_code']}' has no description")
            
            # Check hierarchy consistency
            if acc['detailed_code']:
                # If detailed code exists, minor must exist
                if not acc['minor_code']:
                    errors.append(
                        f"Row {row}: Detailed code '{acc['detailed_code']}' "
                        f"exists but minor code is missing"
                    )
                
                # If minor exists, major must exist
                if acc['minor_code'] and not acc['major_code']:
                    errors.append(
                        f"Row {row}: Minor code '{acc['minor_code']}' "
                        f"exists but major code is missing"
                    )
        
        return errors

    def _build_hierarchy(self, accounts_data):
        """Build hierarchical structure from flat data"""
        
        hierarchy = {
            'majors': {},  # Major heads
            'minors': {},  # Minor heads
            'detailed': {},  # Detailed heads
            'subheads': {}  # Sub-heads (e.g., C03880-10)
        }
        
        for acc in accounts_data:
            # Add major head
            if acc['major_code'] and acc['major_code'] not in hierarchy['majors']:
                hierarchy['majors'][acc['major_code']] = {
                    'code': acc['major_code'],
                    'description': acc['major_desc'],
                    'account_type': self._determine_account_type(acc['major_code']),
                    'minors': []
                }
            
            # Add minor head
            if acc['minor_code'] and acc['minor_code'] not in hierarchy['minors']:
                hierarchy['minors'][acc['minor_code']] = {
                    'code': acc['minor_code'],
                    'description': acc['minor_desc'],
                    'parent_code': acc['major_code'],
                    'account_type': self._determine_account_type(acc['minor_code']),
                    'detailed': []
                }
                
                # Link to parent major
                if acc['major_code'] in hierarchy['majors']:
                    hierarchy['majors'][acc['major_code']]['minors'].append(acc['minor_code'])
            
            # Add detailed head or sub-head
            if acc['detailed_code']:
                # Check if this is a sub-head (contains hyphen like C03880-10)
                if '-' in acc['detailed_code']:
                    # This is a sub-head
                    parent_code = acc['detailed_code'].split('-')[0]  # e.g., C03880
                    
                    # Ensure parent detailed head exists
                    if parent_code not in hierarchy['detailed']:
                        # Need to find the parent's info from previous rows
                        parent_info = self._find_parent_info(accounts_data, parent_code)
                        if parent_info:
                            hierarchy['detailed'][parent_code] = {
                                'code': parent_code,
                                'description': parent_info['description'],
                                'parent_code': parent_info['minor_code'],
                                'account_type': self._determine_account_type(parent_code),
                                'is_control': True,  # Parent of sub-heads is control
                                'subheads': []
                            }
                            
                            # Link to parent minor
                            if parent_info['minor_code'] in hierarchy['minors']:
                                hierarchy['minors'][parent_info['minor_code']]['detailed'].append(parent_code)
                    
                    # Add the sub-head
                    hierarchy['subheads'][acc['detailed_code']] = {
                        'code': acc['detailed_code'],
                        'description': acc['detailed_desc'],
                        'parent_code': parent_code,
                        'account_type': self._determine_account_type(acc['detailed_code'])
                    }
                    
                    # Link to parent detailed
                    if parent_code in hierarchy['detailed']:
                        hierarchy['detailed'][parent_code]['subheads'].append(acc['detailed_code'])
                
                else:
                    # This is a regular detailed head
                    if acc['detailed_code'] not in hierarchy['detailed']:
                        hierarchy['detailed'][acc['detailed_code']] = {
                            'code': acc['detailed_code'],
                            'description': acc['detailed_desc'],
                            'parent_code': acc['minor_code'],
                            'account_type': self._determine_account_type(acc['detailed_code']),
                            'is_control': False,  # Will be set to True if it has sub-heads
                            'subheads': []
                        }
                        
                        # Link to parent minor
                        if acc['minor_code'] in hierarchy['minors']:
                            hierarchy['minors'][acc['minor_code']]['detailed'].append(acc['detailed_code'])
        
        return hierarchy

    def _find_parent_info(self, accounts_data, parent_code):
        """Find parent detailed head info from accounts data"""
        for acc in accounts_data:
            if acc['detailed_code'] == parent_code:
                return {
                    'code': parent_code,
                    'description': acc['detailed_desc'],
                    'minor_code': acc['minor_code'],
                    'major_code': acc['major_code']
                }
        return None

    def _determine_account_type(self, code):
        """Determine Django account type from code prefix"""
        if not code:
            return 'EXPENDITURE'
        
        prefix = code[0].upper()
        
        mapping = {
            'A': 'EXPENDITURE',  # In TMA NAM, A0x is expenditure
            'B': 'REVENUE',       # B0x is tax revenue
            'C': 'REVENUE',       # C0x is non-tax revenue
            'F': 'ASSET',         # F01 is Cash/Bank
            'G': 'ASSET',         # G11 is Investments
        }
        
        return mapping.get(prefix, 'EXPENDITURE')

    def _display_hierarchy_summary(self, hierarchy):
        """Display summary of parsed hierarchy"""
        self.stdout.write(self.style.SUCCESS(f"✓ Parsed structure:"))
        self.stdout.write(f"  Major Heads: {len(hierarchy['majors'])}")
        self.stdout.write(f"  Minor Heads: {len(hierarchy['minors'])}")
        self.stdout.write(f"  Detailed Heads: {len(hierarchy['detailed'])}")
        self.stdout.write(f"  Sub-Heads: {len(hierarchy['subheads'])}\n")

    def _import_accounts(self, hierarchy):
        """Import accounts to database in correct order"""
        stats = {
            'majors': 0,
            'minors': 0,
            'detailed': 0,
            'subheads': 0
        }
        
        created_accounts = {}  # Cache for created account objects
        
        # Step 1: Create Major Heads
        self.stdout.write('\n--- Creating Major Heads ---')
        for code, data in sorted(hierarchy['majors'].items()):
            account = ChartOfAccounts.objects.create(
                account_code=code,
                description=data['description'],
                account_type=data['account_type'],
                parent_account=None,
                is_control_account=True,
                allow_posting=False,
                is_active=True
            )
            created_accounts[code] = account
            stats['majors'] += 1
            self.stdout.write(f"  ✓ {code} - {data['description']}")
        
        # Step 2: Create Minor Heads
        self.stdout.write('\n--- Creating Minor Heads ---')
        for code, data in sorted(hierarchy['minors'].items()):
            parent = created_accounts.get(data['parent_code'])
            
            account = ChartOfAccounts.objects.create(
                account_code=code,
                description=data['description'],
                account_type=data['account_type'],
                parent_account=parent,
                is_control_account=True,
                allow_posting=False,
                is_active=True
            )
            created_accounts[code] = account
            stats['minors'] += 1
            self.stdout.write(f"  ✓ {code} - {data['description']}")
        
        # Step 3: Create Detailed Heads
        self.stdout.write('\n--- Creating Detailed Heads ---')
        for code, data in sorted(hierarchy['detailed'].items()):
            parent = created_accounts.get(data['parent_code'])
            
            # If this detailed head has sub-heads, it's a control account
            is_control = len(data.get('subheads', [])) > 0
            
            account = ChartOfAccounts.objects.create(
                account_code=code,
                description=data['description'],
                account_type=data['account_type'],
                parent_account=parent,
                is_control_account=is_control,
                allow_posting=not is_control,  # Only allow posting if not control
                is_active=True
            )
            created_accounts[code] = account
            stats['detailed'] += 1
            
            control_marker = '[CONTROL]' if is_control else '[TRANS]'
            self.stdout.write(f"  ✓ {code} - {data['description']} {control_marker}")
        
        # Step 4: Create Sub-Heads
        if hierarchy['subheads']:
            self.stdout.write('\n--- Creating Sub-Heads ---')
            for code, data in sorted(hierarchy['subheads'].items()):
                parent = created_accounts.get(data['parent_code'])
                
                account = ChartOfAccounts.objects.create(
                    account_code=code,
                    description=data['description'],
                    account_type=data['account_type'],
                    parent_account=parent,
                    is_control_account=False,
                    allow_posting=True,
                    is_active=True
                )
                created_accounts[code] = account
                stats['subheads'] += 1
                self.stdout.write(f"  ✓ {code} - {data['description']}")
        
        return stats

    def _preview_import(self, hierarchy):
        """Preview what would be imported (dry run)"""
        self.stdout.write('\n' + '=' * 80)
        self.stdout.write('PREVIEW: The following accounts would be created')
        self.stdout.write('=' * 80 + '\n')
        
        # Show sample major heads
        self.stdout.write('MAJOR HEADS (Control Accounts):')
        for code, data in sorted(list(hierarchy['majors'].items())[:5]):
            self.stdout.write(f"  {code:6} - {data['description']}")
        if len(hierarchy['majors']) > 5:
            self.stdout.write(f"  ... and {len(hierarchy['majors']) - 5} more\n")
        
        # Show sample minor heads
        self.stdout.write('\nMINOR HEADS (Control Accounts):')
        for code, data in sorted(list(hierarchy['minors'].items())[:5]):
            self.stdout.write(f"  {code:6} - {data['description']} (under {data['parent_code']})")
        if len(hierarchy['minors']) > 5:
            self.stdout.write(f"  ... and {len(hierarchy['minors']) - 5} more\n")
        
        # Show sample detailed heads
        self.stdout.write('\nDETAILED HEADS (Transaction/Control):')
        for code, data in sorted(list(hierarchy['detailed'].items())[:10]):
            is_control = len(data.get('subheads', [])) > 0
            marker = '[CONTROL]' if is_control else '[TRANS]'
            self.stdout.write(f"  {code:10} - {data['description'][:50]:50} {marker}")
        if len(hierarchy['detailed']) > 10:
            self.stdout.write(f"  ... and {len(hierarchy['detailed']) - 10} more\n")
        
        # Show all sub-heads (there aren't many)
        if hierarchy['subheads']:
            self.stdout.write('\nSUB-HEADS (Transaction Accounts):')
            for code, data in sorted(hierarchy['subheads'].items()):
                self.stdout.write(f"  {code:15} - {data['description']}")

    def _display_import_stats(self, stats):
        """Display import statistics"""
        self.stdout.write('\n' + '=' * 80)
        self.stdout.write('IMPORT STATISTICS')
        self.stdout.write('=' * 80)
        
        total = sum(stats.values())
        
        self.stdout.write(f"\nTotal Accounts Created: {total}")
        self.stdout.write(f"  Major Heads (Level 1): {stats['majors']}")
        self.stdout.write(f"  Minor Heads (Level 2): {stats['minors']}")
        self.stdout.write(f"  Detailed Heads (Level 3): {stats['detailed']}")
        self.stdout.write(f"  Sub-Heads (Level 4): {stats['subheads']}")
        
        self.stdout.write('\n' + '=' * 80)
        self.stdout.write('Next Steps:')
        self.stdout.write('=' * 80)
        self.stdout.write('\n1. Review imported accounts in Django Admin')
        self.stdout.write('2. Create bank account sub-heads under F01xxx')
        self.stdout.write('3. Set up budget allocations')
        self.stdout.write('4. Configure function codes for detailed heads')
        self.stdout.write('5. Test with sample transactions')
        self.stdout.write('\n')
